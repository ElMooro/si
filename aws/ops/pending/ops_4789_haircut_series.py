"""ops/4789 -- tri-party haircut workbooks -> tidy monthly SERIES.
The two banked NY Fed workbooks (current + May2010-Oct2025 history)
hold p10/median/p90 haircuts, dispersion and concentration by ~13
collateral classes -- as files, not chartable rows. This op parses
sheet 'volume_haircut_concentration' from BOTH (layout introspected
and printed, never assumed), melts to per-(class,stat) monthly series,
merges history+current at the Nov-2025 methodology seam (history
strictly before current's first month), and banks:
  data/warm/nyfed-research/haircuts-series/{slug}.json
  + _index.json (id, title, n, first, last) for the engine's scope.
"""
import io
import json
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "ops"))
import boto3  # noqa: E402
from ops_report import report  # noqa: E402

subprocess.run([sys.executable, "-m", "pip", "install", "-q",
                 "openpyxl"], check=False)
import openpyxl  # noqa: E402

B = "justhodl-dashboard-live"
s3 = boto3.client("s3", region_name="us-east-1")
KEYS = {"hist": "data/warm/nyfed-research/haircuts/"
                 "tri-party-repo_preNov25_history.xlsx",
         "curr": "data/warm/nyfed-research/haircuts/"
                 "tri-party-repo_data_current.xlsx"}
STATS = [("median haircut", "median_haircut"),
          ("10th percentile", "p10_haircut"),
          ("90th percentile", "p90_haircut"),
          ("dispersion", "haircut_dispersion"),
          ("concentration", "top3_concentration"),
          ("collateral value", "collateral_value"),
          ("share", "collateral_share")]


def slugify(x):
    return re.sub(r"[^a-z0-9]+", "-", x.lower()).strip("-")[:70]


def parse_wb(raw, rep, tag):
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    sh = None
    for nm in wb.sheetnames:
        if "haircut" in nm.lower():
            sh = wb[nm]
            break
    if sh is None:
        rep.warn(f"{tag}: no haircut sheet; sheets={wb.sheetnames}")
        return {}
    rows = list(sh.iter_rows(values_only=True))
    hdr_i = None
    for i, r0 in enumerate(rows[:8]):
        vals = [str(c) for c in r0 if c is not None]
        if len(vals) >= 4 and any("haircut" in v.lower() or
                                    "percentile" in v.lower()
                                    for v in vals):
            hdr_i = i
            break
    if hdr_i is None:
        hdr_i = 0
    hdr = ["" if c is None else str(c).strip() for c in rows[hdr_i]]
    rep.log(f"{tag} sheet='{sh.title}' header_row={hdr_i} "
            f"cols={len(hdr)}")
    rep.log(f"{tag} header: " + " | ".join(hdr[:14])[:400])
    # date column = first col parsing as date
    series = {}
    for r0 in rows[hdr_i + 1:]:
        if not r0 or r0[0] is None:
            continue
        d0 = r0[0]
        if isinstance(d0, datetime):
            ds = d0.strftime("%Y-%m-%d")
        else:
            m = re.match(r"(\d{4})-(\d{2})", str(d0))
            if not m:
                continue
            ds = f"{m.group(1)}-{m.group(2)}-01"
        for ci in range(1, len(hdr)):
            h = hdr[ci] if ci < len(hdr) else ""
            v = r0[ci] if ci < len(r0) else None
            if not h or v is None:
                continue
            try:
                fv = float(v)
            except Exception:
                continue
            hl = h.lower()
            stat = next((s for pat, s in STATS if pat in hl), None)
            if stat is None:
                continue
            cls = re.sub(r"(?i)(median|10th|90th|percentile|haircut|"
                          r"dispersion|concentration|collateral|value|"
                          r"share|top ?3|margin)", "", h)
            cls = re.sub(r"[\s:_-]+", " ", cls).strip() or "total"
            sid = f"{slugify(cls)}--{stat}"
            series.setdefault(sid, {"cls": cls, "stat": stat,
                                      "obs": {}})["obs"][ds] = fv
    rep.kv(**{f"{tag}_series": len(series),
               f"{tag}_rows": len(rows) - hdr_i - 1})
    return series


def main():
    with report("4789_haircut_series") as rep:
        rep.heading("ops 4789 -- haircut workbooks -> monthly series")
        parsed = {}
        for tag, key in KEYS.items():
            raw = s3.get_object(Bucket=B, Key=key)["Body"].read()
            parsed[tag] = parse_wb(raw, rep, tag)
        merged = {}
        cur = parsed.get("curr") or {}
        his = parsed.get("hist") or {}
        for sid in set(cur) | set(his):
            c = (cur.get(sid) or {}).get("obs") or {}
            h = (his.get(sid) or {}).get("obs") or {}
            cmin = min(c) if c else "9999"
            obs = {**{d: v for d, v in h.items() if d < cmin}, **c}
            meta = cur.get(sid) or his.get(sid)
            merged[sid] = {"cls": meta["cls"], "stat": meta["stat"],
                            "obs": dict(sorted(obs.items()))}
        idx = []
        for sid, m in sorted(merged.items()):
            if len(m["obs"]) < 6:
                continue
            title = (f"Tri-party haircuts: {m['cls']} — "
                      f"{m['stat'].replace('_', ' ')}")
            s3.put_object(Bucket=B,
                Key=f"data/warm/nyfed-research/haircuts-series/{sid}.json",
                Body=json.dumps({"id": sid, "title": title,
                    "observations": [{"date": d, "value": v}
                                      for d, v in m["obs"].items()]},
                    separators=(",", ":")).encode(),
                ContentType="application/json")
            idx.append({"id": sid, "title": title, "n": len(m["obs"]),
                         "first": min(m["obs"]), "last": max(m["obs"])})
        s3.put_object(Bucket=B,
            Key="data/warm/nyfed-research/haircuts-series/_index.json",
            Body=json.dumps({"built_at":
                              datetime.now(timezone.utc).isoformat(),
                              "series": idx},
                             separators=(",", ":")).encode(),
            ContentType="application/json")
        rep.kv(check="haircut_series_banked", value=len(idx))
        for e in idx[:8]:
            rep.ok(f"  {e['id']}: {e['first']} -> {e['last']} n={e['n']}")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        print("ERROR:\n" + traceback.format_exc(), flush=True)
        sys.exit(1)
    sys.exit(0)
