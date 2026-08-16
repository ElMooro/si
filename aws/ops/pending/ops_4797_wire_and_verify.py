"""ops/4797 -- wire + verify the free-source closures:
 (1) copy TE_API_KEY from justhodl-blackswan-watch into justhodl-repo
     env (value never logged); (2) parse the two newest SFTR xlsx
     (openpyxl, runner-side) and PRINT their true structure -- sheet
     names + first 6 rows -- as ground truth vs the engine's mini
     parser; (3) DTCC deep hunt: regex the chart page's 348KB HTML for
     inline JSON arrays / XHR endpoints; (4) async engine run
     (Event + as_of poll, wait v2.3) then assert: 4 TE daily rows,
     D_BTP_BUND_D present with 40-300bp last, WREPOFOR row, SFTR group
     rows>0."""
import gzip
import json
import re
import subprocess
import sys
import time
import urllib.request
from pathlib import Path
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "ops"))
import boto3  # noqa: E402
from botocore.config import Config  # noqa: E402
from ops_report import report  # noqa: E402
subprocess.run([sys.executable, "-m", "pip", "install", "-q",
                 "openpyxl"], check=False)
import openpyxl  # noqa: E402
import io  # noqa: E402

B = "justhodl-dashboard-live"
UA = {"User-Agent": "Mozilla/5.0 (research; raafouis@gmail.com)"}
s3 = boto3.client("s3", region_name="us-east-1")
lam = boto3.client("lambda", region_name="us-east-1",
                    config=Config(read_timeout=30,
                                   retries={"max_attempts": 1}))


def sread(key):
    raw = s3.get_object(Bucket=B, Key=key)["Body"].read()
    if raw[:2] == b"\x1f\x8b":
        raw = gzip.decompress(raw)
    return json.loads(raw)


def main():
    with report("4797_wire_and_verify") as rep:
        rep.heading("ops 4797 -- TE key wire, SFTR truth, DTCC hunt, "
                     "v2.3 verify")

        rep.section("1. TE key -> justhodl-repo env")
        try:
            src = lam.get_function_configuration(
                FunctionName="justhodl-blackswan-watch")
            k = (src.get("Environment") or {}).get("Variables", {}
                  ).get("TE_API_KEY")
            rep.kv(check="te_key_found", value=bool(k),
                    key_len=len(k or ""))
            if k:
                cur = lam.get_function_configuration(
                    FunctionName="justhodl-repo")
                env = (cur.get("Environment") or {}).get("Variables", {})
                if env.get("TE_API_KEY") != k:
                    env["TE_API_KEY"] = k
                    lam.update_function_configuration(
                        FunctionName="justhodl-repo",
                        Environment={"Variables": env})
                    for _ in range(20):
                        st = lam.get_function_configuration(
                            FunctionName="justhodl-repo")
                        if st.get("LastUpdateStatus") == "Successful":
                            break
                        time.sleep(3)
                rep.ok("TE_API_KEY present in justhodl-repo env")
        except Exception as e:
            rep.warn(f"TE wire: {type(e).__name__}: {str(e)[:120]}")

        rep.section("2. SFTR ground truth (openpyxl)")
        html = urllib.request.urlopen(urllib.request.Request(
            "https://www.icmagroup.org/market-practice-and-regulatory-"
            "policy/repo-and-collateral-markets/market-data/"
            "sftr-public-data/", headers=UA), timeout=45
            ).read().decode("utf-8", "replace")
        links = sorted(set(re.findall(
            r"href=\"([^\"]*SFTR-Public-Data-(?:EU|UK)[^\"]*\.xlsx)\"",
            html)))[-2:]
        for lk in links:
            u = lk if lk.startswith("http") else \
                "https://www.icmagroup.org" + lk
            raw = urllib.request.urlopen(urllib.request.Request(
                u, headers=UA), timeout=60).read()
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            rep.log(f"{u.rsplit('/', 1)[-1]}: sheets={wb.sheetnames}")
            sh = wb[wb.sheetnames[0]]
            for i, r0 in enumerate(sh.iter_rows(values_only=True)):
                if i >= 6:
                    break
                rep.log(f"   row{i}: " + repr(r0)[:260])

        rep.section("3. DTCC inline data hunt")
        try:
            page = urllib.request.urlopen(urllib.request.Request(
                "https://www.dtcc.com/charts/daily-total-us-treasury-"
                "trade-fails", headers=UA), timeout=45
                ).read().decode("utf-8", "replace")
            inline = re.findall(r"(\[\{[^\]]{40,400}?\}\])", page)[:3]
            xhr = sorted(set(re.findall(
                r"[\"'](https?://[^\"']*(?:json|api|data)[^\"']*)[\"']",
                page)))[:8]
            rep.kv(check="dtcc_inline_arrays", value=len(inline))
            for a in inline:
                rep.log("  inline: " + a[:200])
            for x in xhr:
                rep.log("  xhr: " + x[:160])
        except Exception as e:
            rep.warn(f"dtcc: {type(e).__name__}")

        rep.section("4. async v2.3 verify")
        pre = None
        try:
            pre = sread("data/repo.json").get("as_of")
        except Exception:
            pass
        lam.invoke(FunctionName="justhodl-repo",
                    InvocationType="Event", Payload=b"{}")
        d = None
        t0 = time.time()
        while time.time() - t0 < 13 * 60:
            time.sleep(25)
            try:
                cur = sread("data/repo.json")
            except Exception:
                continue
            if cur.get("as_of") != pre:
                d = cur
                break
        rep.kv(check="engine_completed", value=bool(d),
                waited_s=round(time.time() - t0, 1))
        if not d:
            return
        allrows = {s0["id"]: s0 for g in d["groups"]
                    for s0 in g["series"]}
        G = {g["name"]: len(g["series"]) for g in d["groups"]}
        for mid in ("DE10Y_TE", "IT10Y_TE", "D_BTP_BUND_D",
                     "WREPOFOR"):
            s0 = allrows.get(mid)
            if s0:
                rep.ok(f"{mid}: {s0['first']} -> {s0['last']} "
                        f"n={s0['n_obs']} last={s0['last_value']}")
            else:
                rep.warn(f"{mid}: ABSENT")
        btp = allrows.get("D_BTP_BUND_D") or {}
        lv = btp.get("last_value")
        rep.kv(check="daily_btp_bund_sane",
                value=bool(lv and 40 <= lv <= 300), last_bp=lv)
        rep.kv(check="sftr_rows",
                value=G.get("SFTR EU/UK repo (ICMA weekly)", 0))
        rep.kv(check="board_total", value=sum(G.values()))


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        print("ERROR:\n" + traceback.format_exc(), flush=True)
        sys.exit(1)
    sys.exit(0)
