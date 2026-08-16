"""ops/4794 -- haircut parser v3. The workbook is LONG format: one row
per (date YYYYMMDD, class#, class name) + positional stat columns; the
header hides sparse above the data band. v3: (a) dump rows 3..13
verbatim; (b) find the header row by KEYWORDS ('collateral','median',
'percentile','concentration','share','value','dispersion','number');
merge with the row below if that adds labels; (c) melt long-format:
date col 0, class name = first string col, stats = header-mapped
numeric cols. Banks ONLY when >=3 stat keywords matched -- never under
guessed labels."""
import io
import json
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "ops"))
import boto3  # noqa: E402
from ops_report import report  # noqa: E402
subprocess.run([sys.executable, "-m", "pip", "install", "-q",
                 "openpyxl"], check=False)
import openpyxl  # noqa: E402

B = "justhodl-dashboard-live"
s3 = boto3.client("s3", region_name="us-east-1")
KEYS = {"hist": "data/warm/nyfed-research/haircuts/"
                 "tri-party-repo_preNov25_history.xlsx",
         "curr": "data/warm/nyfed-research/haircuts/"
                 "tri-party-repo_data_current.xlsx"}
KW = ("collateral", "median", "percentile", "concentr", "share",
       "value", "dispers", "number", "margin", "haircut")
STATMAP = [("median", "median_haircut"), ("10th", "p10_haircut"),
            ("90th", "p90_haircut"), ("dispers", "haircut_dispersion"),
            ("concentr", "top3_concentration"),
            ("share", "collateral_share"),
            ("value", "collateral_value"),
            ("number", "n_observations")]


def slugify(x):
    return re.sub(r"[^a-z0-9]+", "-", x.lower()).strip("-")[:70]


def to_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    sv = str(v).strip()
    m = re.match(r"^(\d{4})(\d{2})(\d{2})(\.0)?$", sv)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", sv)
    return m.group(0) if m else None


def parse_wb(raw, rep, tag):
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    sh = next((wb[n] for n in wb.sheetnames
                if "haircut" in n.lower()), wb[wb.sheetnames[0]])
    rows = list(sh.iter_rows(values_only=True))
    for i in range(3, min(14, len(rows))):
        rep.log(f"{tag} row{i}: " + repr(rows[i])[:340])
    hdr_i = None
    for i in range(0, min(16, len(rows))):
        joined = " ".join(str(c).lower() for c in rows[i] if c)
        if sum(1 for k in KW if k in joined) >= 3:
            hdr_i = i
            break
    if hdr_i is None:
        rep.warn(f"{tag}: NO keyword header found -- refusing to bank "
                 "under guessed labels")
        return {}
    hdr = ["" if c is None else str(c).strip() for c in rows[hdr_i]]
    nxt = rows[hdr_i + 1] if hdr_i + 1 < len(rows) else ()
    if sum(1 for c in nxt if isinstance(c, str) and c.strip()) >= 2 \
            and to_date(nxt[0] if nxt else None) is None:
        hdr = [((hdr[i] + " " + str(nxt[i]).strip()).strip()
                 if i < len(nxt) and isinstance(nxt[i], str)
                 else hdr[i]) for i in range(len(hdr))]
        data_start = hdr_i + 2
    else:
        data_start = hdr_i + 1
    rep.log(f"{tag} HEADER @row{hdr_i}: " + " | ".join(hdr)[:420])
    colstat = {}
    for ci, h in enumerate(hdr):
        hl = h.lower()
        st = next((s for pat, s in STATMAP if pat in hl), None)
        if st:
            colstat[ci] = st
    rep.kv(**{f"{tag}_stat_cols": len(colstat)})
    series = {}
    for r0 in rows[data_start:]:
        if not r0:
            continue
        ds = to_date(r0[0])
        if not ds:
            continue
        cls = next((str(c).strip() for c in r0[1:4]
                     if isinstance(c, str) and c.strip()), None)
        if not cls:
            continue
        for ci, st in colstat.items():
            v = r0[ci] if ci < len(r0) else None
            try:
                fv = float(v)
            except Exception:
                continue
            sid = f"{slugify(cls)}--{st}"
            series.setdefault(sid, {"cls": cls, "stat": st,
                                      "obs": {}})["obs"][ds] = fv
    rep.kv(**{f"{tag}_series": len(series)})
    return series


def main():
    with report("4794_haircut_v3") as rep:
        rep.heading("ops 4794 -- haircut parser v3 (keyword header, "
                     "long format)")
        parsed = {t: parse_wb(s3.get_object(Bucket=B, Key=k)["Body"]
                                .read(), rep, t)
                   for t, k in KEYS.items()}
        cur, his = parsed["curr"], parsed["hist"]
        idx = []
        for sid in sorted(set(cur) | set(his)):
            c = (cur.get(sid) or {}).get("obs") or {}
            h = (his.get(sid) or {}).get("obs") or {}
            cmin = min(c) if c else "9999"
            obs = dict(sorted({**{d: v for d, v in h.items()
                                    if d < cmin}, **c}.items()))
            if len(obs) < 6:
                continue
            meta = cur.get(sid) or his.get(sid)
            title = (f"Tri-party haircuts: {meta['cls']} — "
                      f"{meta['stat'].replace('_', ' ')}")
            s3.put_object(Bucket=B,
                Key="data/warm/nyfed-research/haircuts-series/"
                     f"{sid}.json",
                Body=json.dumps({"id": sid, "title": title,
                    "observations": [{"date": d, "value": v}
                                      for d, v in obs.items()]},
                    separators=(",", ":")).encode(),
                ContentType="application/json")
            idx.append({"id": sid, "title": title, "n": len(obs),
                         "first": min(obs), "last": max(obs)})
        s3.put_object(Bucket=B,
            Key="data/warm/nyfed-research/haircuts-series/_index.json",
            Body=json.dumps({"built_at":
                              datetime.now(timezone.utc).isoformat(),
                              "series": idx},
                             separators=(",", ":")).encode(),
            ContentType="application/json")
        rep.kv(check="haircut_series_banked", value=len(idx))
        for e in idx[:12]:
            rep.ok(f"  {e['id']}: {e['first']} -> {e['last']} n={e['n']}")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        print("ERROR:\n" + traceback.format_exc(), flush=True)
        sys.exit(1)
    sys.exit(0)
