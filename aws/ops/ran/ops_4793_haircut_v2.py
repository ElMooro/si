"""ops/4793 -- haircut parser v2. v1 banked 0: row 0 is blank (title
band) and headers span TWO rows. v2: dump the first 6 raw rows per
workbook verbatim, pick the header band as the two consecutive rows
with the most non-empty cells in rows 0..14, build column names by
concatenating them, then melt to (class,stat) monthly series exactly
as before. Same outputs: haircuts-series/{slug}.json + _index.json."""
import io
import json
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "ops"))
import boto3  # noqa: E402
from ops_report import report  # noqa: E402
subprocess.run([sys.executable, "-m", "pip", "install", "-q",
                 "openpyxl"], check=False)
import openpyxl  # noqa: E402

B = "justhodl-dashboard-live"
s3 = boto3.client("s3", region_name="us-east-1")
KEYS = {"hist": "data/warm/nyfed-research/haircuts/"
                 "tri-party-repo_preNov25_history.xlsx",
         "curr": "data/warm/nyfed-research/haircuts/"
                 "tri-party-repo_data_current.xlsx"}
STATS = [("median", "median_haircut"),
          ("10th", "p10_haircut"), ("90th", "p90_haircut"),
          ("dispers", "haircut_dispersion"),
          ("concentr", "top3_concentration"),
          ("value", "collateral_value"), ("share", "collateral_share")]


def slugify(x):
    return re.sub(r"[^a-z0-9]+", "-", x.lower()).strip("-")[:70]


def parse_wb(raw, rep, tag):
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    sh = next((wb[n] for n in wb.sheetnames
                if "haircut" in n.lower()), wb[wb.sheetnames[0]])
    rows = list(sh.iter_rows(values_only=True))
    for i in range(min(6, len(rows))):
        rep.log(f"{tag} row{i}: " + repr(rows[i][:10])[:300])
    dens = [(sum(1 for c in r0 if c not in (None, "")), i)
             for i, r0 in enumerate(rows[:15])]
    dens.sort(reverse=True)
    h1 = min(dens[0][1], dens[1][1]) if len(dens) > 1 else dens[0][1]
    h2 = h1 + 1 if h1 + 1 < len(rows) else h1
    ncol = max(len(rows[h1]), len(rows[h2]))
    hdr = []
    lastg = ""
    for ci in range(ncol):
        a = rows[h1][ci] if ci < len(rows[h1]) else None
        b0 = rows[h2][ci] if ci < len(rows[h2]) else None
        a = "" if a is None else str(a).strip()
        b0 = "" if b0 is None else str(b0).strip()
        if a:
            lastg = a
        hdr.append((lastg + " " + b0).strip())
    rep.log(f"{tag} header(2-row @{h1},{h2}): "
            + " | ".join(hdr[:12])[:400])
    series = {}
    for r0 in rows[h2 + 1:]:
        if not r0 or r0[0] is None:
            continue
        d0 = r0[0]
        if isinstance(d0, datetime):
            ds = d0.strftime("%Y-%m-%d")
        else:
            m = re.match(r"(\d{4})[-/](\d{1,2})", str(d0))
            if not m:
                continue
            ds = f"{m.group(1)}-{int(m.group(2)):02d}-01"
        for ci in range(1, ncol):
            h = hdr[ci] if ci < len(hdr) else ""
            v = r0[ci] if ci < len(r0) else None
            if not h or v is None:
                continue
            try:
                fv = float(v)
            except Exception:
                continue
            hl = h.lower()
            stat = next((s for pat, s in STATS if pat in hl), None)
            if stat is None:
                continue
            cls = re.sub(r"(?i)(median|10th|90th|percentile|haircut|"
                          r"margin|dispers\w*|concentr\w*|collateral|"
                          r"value|share|top ?3)", "", h)
            cls = re.sub(r"[\s:_()-]+", " ", cls).strip() or "total"
            sid = f"{slugify(cls)}--{stat}"
            series.setdefault(sid, {"cls": cls, "stat": stat,
                                      "obs": {}})["obs"][ds] = fv
    rep.kv(**{f"{tag}_series": len(series)})
    return series


def main():
    with report("4793_haircut_v2") as rep:
        rep.heading("ops 4793 -- haircut parser v2 (2-row headers)")
        parsed = {t: parse_wb(s3.get_object(Bucket=B, Key=k)["Body"]
                                .read(), rep, t)
                   for t, k in KEYS.items()}
        cur, his = parsed["curr"], parsed["hist"]
        idx = []
        for sid in sorted(set(cur) | set(his)):
            c = (cur.get(sid) or {}).get("obs") or {}
            h = (his.get(sid) or {}).get("obs") or {}
            cmin = min(c) if c else "9999"
            obs = dict(sorted({**{d: v for d, v in h.items()
                                    if d < cmin}, **c}.items()))
            if len(obs) < 6:
                continue
            meta = cur.get(sid) or his.get(sid)
            title = (f"Tri-party haircuts: {meta['cls']} — "
                      f"{meta['stat'].replace('_', ' ')}")
            s3.put_object(Bucket=B,
                Key="data/warm/nyfed-research/haircuts-series/"
                     f"{sid}.json",
                Body=json.dumps({"id": sid, "title": title,
                    "observations": [{"date": d, "value": v}
                                      for d, v in obs.items()]},
                    separators=(",", ":")).encode(),
                ContentType="application/json")
            idx.append({"id": sid, "title": title, "n": len(obs),
                         "first": min(obs), "last": max(obs)})
        s3.put_object(Bucket=B,
            Key="data/warm/nyfed-research/haircuts-series/_index.json",
            Body=json.dumps({"built_at":
                              datetime.now(timezone.utc).isoformat(),
                              "series": idx},
                             separators=(",", ":")).encode(),
            ContentType="application/json")
        rep.kv(check="haircut_series_banked", value=len(idx))
        for e in idx[:10]:
            rep.ok(f"  {e['id']}: {e['first']} -> {e['last']} n={e['n']}")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        print("ERROR:\n" + traceback.format_exc(), flush=True)
        sys.exit(1)
    sys.exit(0)
